"""
services/financiero.py — Lógica de negocio financiera + becas.

Reglas críticas:
- Toda escritura de pagos corre bajo transacción atómica (session.begin()).
- Los pagos son INMUTABLES; ningún endpoint PUT/DELETE expuesto.
- Correcciones = nuevo pago con es_ajuste=True.
- Bloqueo inscripción: verificar cuotas vencidas vs umbral de carrera.
  EXCEPCIÓN: alumno con beca 100% no se bloquea (override implícito).
- Multi-beca: aplica el MAYOR porcentaje entre las activas y vigentes.
"""

from __future__ import annotations

import io
from datetime import date
from decimal import Decimal
from typing import Any, List, Optional, cast

from dateutil.relativedelta import relativedelta
from sqlalchemy.orm import Session
from sqlalchemy import func

from app.models.financiero import (
    BecaActiva,
    BecaCatalogo,
    ConceptoArancel,
    Cuota,
    FuenteBeca,
    Pago,
    AuditoriaOverrideMora,
)
from app.models.users import User
from app.models.carrera import Carrera
from app.schemas.financiero import (
    BecaActivaOut,
    CuotaOut,
    CuotaVencidaDetalle,
    EstadoDeudaOut,
)


# ═══════════════════════════════════════════════════════════════════════
# BECAS — helpers
# ═══════════════════════════════════════════════════════════════════════


def _becas_vigentes(alumno_id: int, db: Session) -> List[BecaActiva]:
    """Retorna todas las BecaActivas con estado_renovacion='vigente' del alumno."""
    return (
        db.query(BecaActiva)
        .filter(
            BecaActiva.alumno_id == alumno_id,
            BecaActiva.estado_renovacion == "vigente",
        )
        .all()
    )


def calcular_descuento_beca(
    alumno_id: int, db: Session
) -> tuple[Decimal, Optional[BecaActiva]]:
    """
    Retorna (porcentaje_descuento, beca_activa_aplicada).
    Aplica el MAYOR porcentaje entre todas las becas vigentes.
    Si no tiene beca: (0, None).
    """
    becas = _becas_vigentes(alumno_id, db)
    if not becas:
        return Decimal("0"), None
    mejor = max(
        becas, key=lambda b: b.beca.porcentaje_descuento if b.beca else Decimal("0")
    )
    return mejor.beca.porcentaje_descuento if mejor.beca else Decimal("0"), mejor


def tiene_beca_100(alumno_id: int, db: Session) -> bool:
    """True si el alumno tiene al menos una beca vigente con 100% de descuento."""
    becas = _becas_vigentes(alumno_id, db)
    return any(b.beca and b.beca.porcentaje_descuento >= Decimal("100") for b in becas)


def get_becas_activas_out(alumno_id: int, db: Session) -> List[BecaActivaOut]:
    """Shape exacto exigido por el enunciado para GET /becas/alumno/{id}/activas."""
    becas = db.query(BecaActiva).filter(BecaActiva.alumno_id == alumno_id).all()
    result = []
    for ba in becas:
        beca = ba.beca
        fuente = ba.fuente
        result.append(
        BecaActivaOut(
            id=cast(int, ba.id),
            beca_nombre=str(beca.nombre) if beca else "—",
            fuente=str(fuente.nombre) if fuente else "—",
            es_externa=bool(fuente.es_externa) if fuente else False,
            porcentaje_descuento=beca.porcentaje_descuento
            if beca
            else Decimal("0"),
            periodo_inicio=str(ba.periodo_inicio),
            periodo_fin=str(ba.periodo_fin) if ba.periodo_fin else None,
            promedio_minimo_requerido=cast(Decimal, ba.promedio_minimo_requerido),
            promedio_actual=cast(Decimal, ba.promedio_actual),
            estado_renovacion=str(ba.estado_renovacion),
        )
        )
    return result


# ═══════════════════════════════════════════════════════════════════════
# CUOTAS — generación
# ═══════════════════════════════════════════════════════════════════════


def generar_cuotas_alumno(
    alumno_id: int,
    concepto_id: int,
    periodos: List[str],
    fecha_vencimiento_base: date,
    generado_por: int,
    db: Session,
) -> List[Cuota]:
    """
    Genera cuotas para un alumno dado un concepto y lista de períodos.
    Descuenta el mayor porcentaje de beca vigente.
    Toda la operación es atómica.
    """
    concepto = (
        db.query(ConceptoArancel).filter(ConceptoArancel.id == concepto_id).first()
    )
    if not concepto:
        raise ValueError(f"Concepto arancel {concepto_id} no existe")

    porcentaje, beca_activa = calcular_descuento_beca(alumno_id, db)
    descuento_factor = porcentaje / Decimal("100")

    nuevas: List[Cuota] = []
    for i, periodo in enumerate(periodos):
        fecha_venc = fecha_vencimiento_base + relativedelta(months=i)
        monto_desc = (concepto.monto_base * descuento_factor).quantize(Decimal("0.01"))
        cuota = Cuota(
            alumno_id=alumno_id,
            concepto_id=concepto_id,
            periodo=periodo,
            monto=concepto.monto_base,
            monto_descuento=monto_desc,
            fecha_vencimiento=fecha_venc,
            estado="pendiente",
            beca_aplicada_id=beca_activa.id if beca_activa else None,
            generado_por=generado_por,
        )
        db.add(cuota)
        nuevas.append(cuota)

    db.flush()
    return nuevas


def cuota_to_out(cuota: Cuota) -> CuotaOut:
    beca_nombre = None
    fuente_beca = None
    es_beca_externa = None
    if cuota.beca_aplicada:
        beca_nombre = (
            cuota.beca_aplicada.beca.nombre if cuota.beca_aplicada.beca else None
        )
        fuente_beca = (
            cuota.beca_aplicada.fuente.nombre if cuota.beca_aplicada.fuente else None
        )
        es_beca_externa = (
            cuota.beca_aplicada.fuente.es_externa
            if cuota.beca_aplicada.fuente
            else None
        )

    pago_id = None
    comprobante_estado = None
    comprobante_url_pdf = None
    pagos_list: list = list(cast(list[Any], cuota.pagos or []))
    if pagos_list:
        ultimo_pago = max(pagos_list, key=lambda p: p.fecha_pago)
        pago_id = ultimo_pago.id
        if ultimo_pago.comprobante:
            comprobante_estado = ultimo_pago.comprobante.estado_emision
            comprobante_url_pdf = ultimo_pago.comprobante.url_pdf

    return CuotaOut(
        id=cast(int, cuota.id),
        alumno_id=cast(int, cuota.alumno_id),
        concepto_id=cast(int, cuota.concepto_id),
        periodo=str(cuota.periodo),
        monto=cast(Decimal, cuota.monto),
        monto_descuento=cast(Decimal, cuota.monto_descuento),
        monto_a_pagar=cast(Decimal, cuota.monto - cuota.monto_descuento),
        fecha_vencimiento=cast(date, cuota.fecha_vencimiento),
        estado=str(cuota.estado),
        beca_nombre=beca_nombre,
        fuente_beca=fuente_beca,
        es_beca_externa=es_beca_externa,
        pago_id=pago_id,
        comprobante_estado=comprobante_estado,
        comprobante_url_pdf=comprobante_url_pdf,
    )


# ═══════════════════════════════════════════════════════════════════════
# PAGOS — registro atómico
# ═══════════════════════════════════════════════════════════════════════


def registrar_pago(
    cuota_id: int,
    monto_pagado: Decimal,
    metodo: str,
    registrado_por: int,
    db: Session,
    referencia: Optional[str] = None,
    pago_ajuste_ref_id: Optional[int] = None,
    nota_ajuste: Optional[str] = None,
) -> Pago:
    """
    Registra un pago de forma atómica.
    Soporta pagos parciales — la cuota se marca 'pagada' solo cuando el
    total pagado >= monto_a_pagar.
    Los pagos son INMUTABLES una vez creados.
    """
    cuota = db.query(Cuota).filter(Cuota.id == cuota_id).first()
    if not cuota:
        raise ValueError(f"Cuota {cuota_id} no encontrada")
    if cuota.estado == "anulada":
        raise ValueError("No se puede pagar una cuota anulada")

    es_ajuste = pago_ajuste_ref_id is not None

    pago = Pago(
        cuota_id=cuota_id,
        monto_pagado=monto_pagado,
        metodo=metodo,
        referencia=referencia,
        registrado_por=registrado_por,
        pago_ajuste_ref_id=pago_ajuste_ref_id,
        es_ajuste=es_ajuste,
        nota_ajuste=nota_ajuste,
    )
    db.add(pago)
    db.flush()

    # Recalcular total pagado vs monto a pagar
    total_pagado = (
        db.query(func.sum(Pago.monto_pagado))
        .filter(Pago.cuota_id == cuota_id, Pago.es_ajuste == False)  # noqa: E712
        .scalar()
        or Decimal("0")
    )
    monto_a_pagar = cuota.monto - cuota.monto_descuento
    if total_pagado >= monto_a_pagar:
        setattr(cuota, 'estado', "pagada")
        db.flush()

    return pago


# ═══════════════════════════════════════════════════════════════════════
# BLOQUEO INSCRIPCIÓN POR MORA
# ═══════════════════════════════════════════════════════════════════════


def verificar_deuda_inscripcion(
    alumno_id: int,
    db: Session,
    es_admin: bool = False,
) -> EstadoDeudaOut:
    """
    Verifica si el alumno puede inscribirse.

    EXCEPCIÓN CRÍTICA: si tiene beca 100%, no se bloquea.
    Override disponible solo para administradores.
    """
    hoy = date.today()

    # obtener carrera del alumno para saber umbral
    alumno = db.query(User).filter(User.id == alumno_id).first()
    max_mora = 1  # default
    if alumno and alumno.carrera_id:
        carrera = db.query(Carrera).filter(Carrera.id == alumno.carrera_id).first()
        if carrera:
            max_mora = carrera.max_cuotas_mora

    cuotas_vencidas = (
        db.query(Cuota)
        .filter(
            Cuota.alumno_id == alumno_id,
            Cuota.estado == "pendiente",
            Cuota.fecha_vencimiento < hoy,
        )
        .order_by(Cuota.fecha_vencimiento.asc())
        .all()
    )

    detalle = [
        CuotaVencidaDetalle(
            cuota_id=cast(int, c.id),
            periodo=str(c.periodo),
            monto_a_pagar=cast(Decimal, c.monto - c.monto_descuento),
            fecha_vencimiento=cast(date, c.fecha_vencimiento),
            dias_vencida=(hoy - c.fecha_vencimiento).days,
        )
        for c in cuotas_vencidas
    ]

    _tiene_beca_100 = tiene_beca_100(alumno_id, db)
    bloqueado = len(cuotas_vencidas) >= max_mora and not _tiene_beca_100

    return EstadoDeudaOut(
        bloqueado=bool(bloqueado),
        cuotas_vencidas=len(cuotas_vencidas),
        max_permitidas=cast(int, max_mora),
        detalle=detalle,
        tiene_beca_100=_tiene_beca_100,
        override_disponible=es_admin,
    )


def registrar_override_mora(
    alumno_id: int,
    admin_id: int,
    db: Session,
    oferta_materia_id: Optional[int] = None,
    motivo: Optional[str] = None,
) -> AuditoriaOverrideMora:
    """Registra en auditoría que un admin omitió el bloqueo por mora."""
    registro = AuditoriaOverrideMora(
        alumno_id=alumno_id,
        admin_id=admin_id,
        oferta_materia_id=oferta_materia_id,
        motivo=motivo,
    )
    db.add(registro)
    db.flush()
    return registro


# ═══════════════════════════════════════════════════════════════════════
# REPORTE RENDICIÓN — Excel
# ═══════════════════════════════════════════════════════════════════════


def export_rendicion_excel(
    fuente_id: int,
    periodo: Optional[str],
    db: Session,
) -> bytes:
    """
    Genera bytes de un archivo Excel con la rendición de becas por fuente.
    Columnas: Alumno, Cédula, Carrera, Beca, Fuente, % Descuento, Monto Becado, Periodo.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise RuntimeError("openpyxl no instalado — ejecutar: pip install openpyxl")

    query = (
        db.query(BecaActiva, User, Carrera, BecaCatalogo, FuenteBeca)
        .join(User, BecaActiva.alumno_id == User.id)
        .join(BecaCatalogo, BecaActiva.beca_id == BecaCatalogo.id)
        .join(FuenteBeca, BecaActiva.fuente_id == FuenteBeca.id)
        .outerjoin(Carrera, User.carrera_id == Carrera.id)
        .filter(BecaActiva.fuente_id == fuente_id)
    )
    if periodo:
        query = query.filter(
            BecaActiva.periodo_inicio <= periodo,
            (BecaActiva.periodo_fin >= periodo) | (BecaActiva.periodo_fin.is_(None)),
        )

    rows = query.all()

    beca_activa_ids = [ba.id for ba, *_ in rows]
    montos_becados: dict[int, Decimal] = {}
    if beca_activa_ids:
        montos_q = (
            db.query(Cuota.beca_aplicada_id, func.sum(Cuota.monto_descuento))
            .filter(Cuota.beca_aplicada_id.in_(beca_activa_ids))
        )
        if periodo:
            montos_q = montos_q.filter(Cuota.periodo == periodo)
        montos_becados = {r[0]: r[1] for r in montos_q.group_by(Cuota.beca_aplicada_id).all() if r[0] is not None}

    wb = openpyxl.Workbook()
    ws = wb.active
    # AUDIT-FIX B-8: sheet puede ser None en openpyxl
    if ws is None:
        raise ValueError("No se encontró el worksheet activo en el Excel")
    ws.title = "Rendición Becas"

    # Header
    headers = [
        "Alumno",
        "Cédula",
        "Carrera",
        "Beca",
        "Fuente",
        "% Descuento",
        "Monto Becado (Gs.)",
        "Período",
    ]
    header_fill = PatternFill("solid", fgColor="1E40AF")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data
    for row_idx, (ba, user, carrera, beca, fuente) in enumerate(rows, 2):
        monto_becado = montos_becados.get(ba.id) or Decimal("0")

        ws.cell(row=row_idx, column=1, value=user.nombre or user.username)
        ws.cell(
            row=row_idx, column=2, value=user.username
        )  # cedula = username por ahora
        ws.cell(row=row_idx, column=3, value=carrera.nombre if carrera else "—")
        ws.cell(row=row_idx, column=4, value=beca.nombre)
        ws.cell(row=row_idx, column=5, value=fuente.nombre)
        ws.cell(row=row_idx, column=6, value=float(beca.porcentaje_descuento))
        ws.cell(row=row_idx, column=7, value=float(monto_becado))
        ws.cell(row=row_idx, column=8, value=periodo or ba.periodo_inicio)

    # Auto-width
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        # AUDIT-FIX: col[0] puede ser MergedCell que no tiene column_letter
        cell_col_idx: int | None = getattr(col[0], 'column', None)
        if cell_col_idx is not None:
            col_letter = get_column_letter(cell_col_idx)
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
